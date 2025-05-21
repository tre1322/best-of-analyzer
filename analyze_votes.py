import pandas as pd
from collections import Counter
from rapidfuzz import fuzz, process
from openpyxl import Workbook
import os
import re
from datetime import datetime, timedelta

def load_business_master(filepath="business_master.csv"):
    df = pd.read_csv(filepath)
    df = df.dropna(subset=['Business Name'])
    df['normalized'] = df['Business Name'].apply(normalize_name)
    return df

def match_to_master(raw_name, business_df, category=None, threshold=85):
    norm = normalize_name(raw_name)

    if category and 'Category' in business_df.columns:
        filtered = business_df[business_df['Category'].str.lower().str.contains(category.lower(), na=False)]
    else:
        filtered = business_df

    matches = filtered['normalized'].tolist()
    match, score, _ = process.extractOne(norm, matches, scorer=fuzz.token_sort_ratio)
    if score >= threshold:
        return filtered[filtered['normalized'] == match].iloc[0]['Business Name']
    return raw_name

def load_anchors(filepath="anchors.csv"):
    try:
        df = pd.read_csv(filepath)
        return {row['anchor'].strip().lower(): row['canonical'].strip() for _, row in df.iterrows()}
    except Exception:
        return {}

def normalize_name(name):
    name = str(name).lower()
    name = re.sub(r"[’'`]", "", name)
    name = re.sub(r"[.-]", " ", name)
    name = name.replace("&", "and")
    name = re.sub(r"[^a-z0-9 ]", "", name)
    name = re.sub(r"\s+", " ", name)
    return name.strip()

def get_best_anchor(norm, anchors, threshold=80):
    for anchor, canonical in anchors.items():
        if anchor in norm:
            return canonical
    match, score, _ = process.extractOne(norm, anchors.keys(), scorer=fuzz.partial_ratio)
    if match and score >= threshold:
        return anchors[match]
    return None

def prefix_grouping(raw_name, canonical_map, existing_canonicals, min_prefix=4):
    norm = normalize_name(raw_name)
    prefix = norm[:min_prefix]
    for canon in existing_canonicals:
        if normalize_name(canon).startswith(prefix):
            return canon
    return raw_name

def deduplicate_canonicals(canonicals, threshold=90):
    final_map = {}
    buckets = []

    for canon in canonicals:
        matched = False
        for bucket in buckets:
            if fuzz.token_sort_ratio(canon, bucket[0]) >= threshold:
                bucket.append(canon)
                final_map[canon] = bucket[0]
                matched = True
                break
        if not matched:
            buckets.append([canon])
            final_map[canon] = canon
    return final_map

def build_canonical_map(vote_names, anchors, business_df, category_name=None):
    canonical_map = {}
    raw_to_canon = {}
    existing_canonicals = set()

    for raw in vote_names:
        norm = normalize_name(raw)
        canonical = get_best_anchor(norm, anchors)
        if not canonical:
            canonical = match_to_master(raw, business_df, category=category_name)
        if not canonical or canonical == raw:
            canonical = prefix_grouping(raw, canonical_map, existing_canonicals)

        raw_to_canon[raw] = canonical
        existing_canonicals.add(canonical)

    deduped = deduplicate_canonicals(set(raw_to_canon.values()))
    for raw, canon in raw_to_canon.items():
        canonical_map[raw] = deduped[canon]
    return canonical_map

def detect_time_clustered_votes(vote_df, time_col="timestamp", business_col="canonical", time_window_minutes=10, cluster_threshold=8):
    fraud_indexes = {}
    vote_df = vote_df.copy()

    vote_df['parsed_time'] = pd.to_datetime(vote_df[time_col], errors='coerce')
    vote_df = vote_df.sort_values(by='parsed_time')
    grouped = vote_df.groupby(business_col)

    for biz, group in grouped:
        times = group['parsed_time'].tolist()
        indices = group.index.tolist()

        flagged = set()
        for i in range(len(times)):
            if indices[i] in flagged:
                continue

            cluster = [indices[i]]
            j = i + 1
            while j < len(times) and (times[j] - times[i]) <= timedelta(minutes=time_window_minutes):
                cluster.append(indices[j])
                j += 1

            if len(cluster) >= cluster_threshold:
                for k in cluster:
                    fraud_indexes[k] = f"Rule 6: {len(cluster)} votes for '{biz}' within {time_window_minutes} minutes"
                    flagged.add(k)

    return fraud_indexes


def analyze_votes(file_path, category_name, output_path='final_results.xlsx'):
    os.makedirs('uploads', exist_ok=True)
    if os.path.exists(output_path):
        os.remove(output_path)

    df = pd.read_excel(file_path, engine='openpyxl', header=0)
    df.columns = df.columns.str.strip().str.lower()
    category_name = category_name.lower()

    if category_name not in df.columns:
        raise ValueError(f"Column '{category_name}' not found in spreadsheet.")

    vote_col = category_name
    ip_col = "ip address"
    time_col = "start date"
    if not ip_col or not time_col:
        raise KeyError("Missing required columns: IP Address or Timestamp.")
    
    df.rename(columns={ip_col: "ip address", time_col: "timestamp"}, inplace=True)

    df['votes filled'] = df.notnull().sum(axis=1)
    fraud_rows = {}

    ip_counts = df["ip address"].value_counts()
    repeat_ips = ip_counts[ip_counts > 1].index.tolist()
    for idx, row in df.iterrows():
        if row["ip address"] in repeat_ips and row["votes filled"] <= 2:
            fraud_rows[idx] = "Duplicate IP with only one vote"

    vote_df = df[[vote_col, "ip address", "timestamp"]].dropna().copy()
    vote_df['raw'] = vote_df[vote_col].astype(str).str.strip()
    anchors = load_anchors()
    business_df = load_business_master()
    canonical_map = build_canonical_map(vote_df['raw'].drop_duplicates(), anchors, business_df, category_name)

    vote_df['canonical'] = vote_df['raw'].map(lambda x: canonical_map.get(x, x))
    repeated = vote_df.groupby(['ip address', 'canonical']).size()
    suspicious = repeated[repeated > 1]
    if not suspicious.empty:
        suspect_set = set(suspicious.index)
        for idx, row in df.iterrows():
            raw = str(row.get(vote_col, "")).strip()
            ip = row["ip address"]
            canonical = canonical_map.get(raw, raw)
            if canonical and (ip, canonical) in suspect_set:
                fraud_rows[idx] = f"Rule 5: Category-wide repeat vote for '{canonical}' from IP {ip}"

    # Rule 6: Detect time-based vote clusters
    time_fraud = detect_time_clustered_votes(vote_df[['canonical', 'timestamp', 'ip address']], business_col='canonical')
    for idx, reason in time_fraud.items():
        fraud_rows[idx] = reason

    legit_df = df.drop(index=fraud_rows.keys())
    values = legit_df[vote_col].dropna().astype(str).str.strip()
    mapped_names = values.map(lambda x: canonical_map.get(x, x))
    vote_counts = Counter(mapped_names)
    sorted_votes = sorted(vote_counts.items(), key=lambda x: x[1], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["Rank", "Business", "Votes"])
    for i, (biz, count) in enumerate(sorted_votes[:3], start=1):
        ws.append([i, biz, count])

    fraud_ws = wb.create_sheet("Fraud Report")
    fraud_ws.append(["Row", "Reason"])
    for idx, reason in sorted(fraud_rows.items()):
        fraud_ws.append([f"A{idx+2}", reason])

    tracker = wb.create_sheet("Vote Tracker")
    tracker.append(["Row #", "IP Address", "Original Vote", "Normalized", "Canonical Name", "Counted?"])
    for idx, row in df.iterrows():
        raw_vote = str(row.get(vote_col, "")).strip()
        if not raw_vote or raw_vote.lower() in ["nan", "none"]:
            continue
        norm = normalize_name(raw_vote)
        canonical = canonical_map.get(raw_vote, raw_vote)
        ip = row.get("ip address", "")
        counted = "No" if idx in fraud_rows else "Yes"
        tracker.append([f"A{idx+2}", ip, raw_vote, norm, canonical, counted])

    wb.save(output_path)
    return output_path